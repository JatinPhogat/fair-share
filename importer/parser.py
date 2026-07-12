import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher

from openpyxl import load_workbook


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None

CANONICAL_NAMES = {}

KNOWN_NAME_VARIANTS = {
    "priya s": "Priya",
    "priya": "Priya",
    "rohan": "Rohan",
    "aisha": "Aisha",
    "meera": "Meera",
    "dev": "Dev",
    "sam": "Sam",
}


def normalize_name(name: str) -> str:
    if not name:
        return ""
    stripped = name.strip()
    key = stripped.lower()
    if key in KNOWN_NAME_VARIANTS:
        return KNOWN_NAME_VARIANTS[key]
    return stripped.title()


def parse_split_details(raw: str) -> dict[str, str]:
    result = {}
    if not raw:
        return result
    parts = raw.split(";")
    for part in parts:
        part = part.strip()
        match = re.match(r"^(.+?)\s+([\d.]+%?)$", part)
        if match:
            name = normalize_name(match.group(1))
            value = match.group(2)
            result[name] = value
    return result


def parse_participants(raw: str) -> list[str]:
    if not raw:
        return []
    return [normalize_name(p) for p in raw.split(";") if p.strip()]


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def read_file(file_obj) -> list[dict]:
    """Read CSV or XLSX file and return list of row dicts."""
    filename = getattr(file_obj, "name", "")

    if filename.endswith((".xlsx", ".xls")):
        return _read_xlsx(file_obj)
    return _read_csv(file_obj)


def _read_xlsx(file_obj) -> list[dict]:
    wb = load_workbook(file_obj, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            row_dict[key] = val
        result.append(row_dict)
    return result


def _read_csv(file_obj) -> list[dict]:
    if isinstance(file_obj, (bytes, bytearray)):
        file_obj = io.StringIO(file_obj.decode("utf-8"))
    elif hasattr(file_obj, "read"):
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode("utf-8")
        file_obj = io.StringIO(content)

    reader = csv.DictReader(file_obj)
    return [dict(row) for row in reader]


def detect_anomalies(rows: list[dict], group_members: list[str], member_dates: dict) -> list[dict]:
    """Run all anomaly detectors on parsed rows.

    Args:
        rows: list of row dicts from read_file
        group_members: list of canonical member names
        member_dates: dict of {name: {"joined": date, "left": date|None}}

    Returns:
        list of dicts with keys: row_number, anomaly_type, description, severity, auto_resolution
    """
    anomalies = []
    seen_expenses = []

    for i, row in enumerate(rows):
        row_num = i + 2  # +2 because row 1 is header, data starts at row 2

        anomalies.extend(_check_missing_fields(row, row_num))
        anomalies.extend(_check_name_issues(row, row_num, group_members))
        anomalies.extend(_check_amount(row, row_num))
        anomalies.extend(_check_currency(row, row_num))
        anomalies.extend(_check_date(row, row_num))
        anomalies.extend(_check_settlement(row, row_num))
        anomalies.extend(_check_split(row, row_num))
        anomalies.extend(_check_membership_temporal(row, row_num, member_dates))
        anomalies.extend(_check_duplicate(row, row_num, seen_expenses))
        anomalies.extend(_check_unknown_participants(row, row_num, group_members))

        seen_expenses.append((row_num, row))

    return anomalies


def _check_missing_fields(row: dict, row_num: int) -> list[dict]:
    results = []
    if not row.get("paid_by") or str(row["paid_by"]).strip() == "" or row["paid_by"] is None:
        results.append({
            "row_number": row_num,
            "anomaly_type": "missing_field",
            "description": f"Missing 'paid_by' field. Notes: {row.get('notes', '')}",
            "severity": "error",
            "auto_resolution": "Requires user to specify who paid.",
        })
    return results


def _check_name_issues(row: dict, row_num: int, group_members: list[str]) -> list[dict]:
    results = []

    paid_by_raw = str(row.get("paid_by", "")).strip()
    if paid_by_raw and paid_by_raw != normalize_name(paid_by_raw):
        results.append({
            "row_number": row_num,
            "anomaly_type": "name_casing",
            "description": f"Payer name '{paid_by_raw}' has inconsistent casing/whitespace.",
            "severity": "info",
            "auto_resolution": f"Normalized to '{normalize_name(paid_by_raw)}'.",
        })

    if paid_by_raw:
        normalized = normalize_name(paid_by_raw)
        if normalized not in group_members:
            for member in group_members:
                if similarity(normalized, member) > 0.7 and normalized != member:
                    results.append({
                        "row_number": row_num,
                        "anomaly_type": "name_variant",
                        "description": f"Payer '{paid_by_raw}' might be '{member}'.",
                        "severity": "warning",
                        "auto_resolution": f"Mapped '{paid_by_raw}' to '{member}'.",
                    })
                    break

    split_with = str(row.get("split_with", ""))
    for name in split_with.split(";"):
        name = name.strip()
        if not name:
            continue
        if name.lower() != normalize_name(name).lower():
            pass  # covered by per-name normalization

    return results


def _check_amount(row: dict, row_num: int) -> list[dict]:
    results = []
    try:
        amount = Decimal(str(row.get("amount", 0)))
    except (InvalidOperation, ValueError):
        results.append({
            "row_number": row_num,
            "anomaly_type": "missing_field",
            "description": f"Invalid amount: '{row.get('amount')}'",
            "severity": "error",
            "auto_resolution": "Requires user correction.",
        })
        return results

    if amount == 0:
        results.append({
            "row_number": row_num,
            "anomaly_type": "zero_amount",
            "description": f"Expense '{row.get('description')}' has zero amount. Notes: {row.get('notes', '')}",
            "severity": "warning",
            "auto_resolution": "Flagged for skip — likely a voided/placeholder entry.",
        })
    elif amount < 0:
        results.append({
            "row_number": row_num,
            "anomaly_type": "negative_amount",
            "description": f"Negative amount {amount} for '{row.get('description')}'. Treating as refund.",
            "severity": "warning",
            "auto_resolution": "Will be imported as a refund (reverse split).",
        })

    return results


def _check_currency(row: dict, row_num: int) -> list[dict]:
    results = []
    currency = row.get("currency")
    if not currency or str(currency).strip().upper() == "NONE" or str(currency).strip() == "":
        results.append({
            "row_number": row_num,
            "anomaly_type": "currency_missing",
            "description": f"Missing currency for '{row.get('description')}'. Notes: {row.get('notes', '')}",
            "severity": "warning",
            "auto_resolution": "Defaulting to INR.",
        })
    elif str(currency).strip().upper() == "USD":
        results.append({
            "row_number": row_num,
            "anomaly_type": "multi_currency",
            "description": f"USD amount for '{row.get('description')}'. Will convert using session exchange rate.",
            "severity": "info",
            "auto_resolution": "Will convert to INR using the configured exchange rate.",
        })

    return results


def _check_date(row: dict, row_num: int) -> list[dict]:
    results = []
    date_val = row.get("date")
    if not date_val:
        return results

    try:
        if isinstance(date_val, str):
            parsed = datetime.strptime(date_val, "%Y-%m-%d").date()
        elif isinstance(date_val, (datetime, date)):
            parsed = date_val if isinstance(date_val, date) else date_val.date()
        else:
            parsed = None
    except ValueError:
        results.append({
            "row_number": row_num,
            "anomaly_type": "date_anomaly",
            "description": f"Cannot parse date '{date_val}'.",
            "severity": "error",
            "auto_resolution": "Requires user correction.",
        })
        return results

    if parsed and parsed.year < 2020:
        results.append({
            "row_number": row_num,
            "anomaly_type": "date_anomaly",
            "description": f"Date {parsed} has year {parsed.year}, likely wrong (expected 2026). '{row.get('description')}'",
            "severity": "error",
            "auto_resolution": f"Suggested correction: {parsed.replace(year=2026)}",
        })

    notes = str(row.get("notes", "")).lower()
    desc = str(row.get("description", "")).lower()
    if "date" in notes or "format" in notes:
        results.append({
            "row_number": row_num,
            "anomaly_type": "date_anomaly",
            "description": f"Note indicates date ambiguity: '{row.get('notes')}'. Date is {parsed}.",
            "severity": "warning",
            "auto_resolution": "Flagged for user review.",
        })

    return results


def _check_settlement(row: dict, row_num: int) -> list[dict]:
    results = []
    split_type = row.get("split_type")
    desc = str(row.get("description", "")).lower()
    notes = str(row.get("notes", "")).lower()

    settlement_keywords = ["paid back", "settlement", "settled", "deposit", "repay"]
    is_settlement_desc = any(kw in desc for kw in settlement_keywords)
    is_settlement_note = "settlement" in notes or "not an expense" in notes

    if split_type is None or str(split_type).strip() == "" or str(split_type).strip().upper() == "NONE":
        if is_settlement_desc or is_settlement_note:
            results.append({
                "row_number": row_num,
                "anomaly_type": "settlement",
                "description": f"'{row.get('description')}' appears to be a settlement, not an expense. No split_type set.",
                "severity": "warning",
                "auto_resolution": "Will be imported as a payment/settlement record.",
            })
        else:
            results.append({
                "row_number": row_num,
                "anomaly_type": "missing_field",
                "description": f"Missing split_type for '{row.get('description')}'.",
                "severity": "error",
                "auto_resolution": "Requires user to specify split type or mark as settlement.",
            })

    deposit_keywords = ["deposit share", "deposit"]
    if any(kw in desc for kw in deposit_keywords):
        participants = parse_participants(str(row.get("split_with", "")))
        if len(participants) == 1:
            results.append({
                "row_number": row_num,
                "anomaly_type": "settlement",
                "description": f"'{row.get('description')}' looks like a direct payment to {participants[0]}.",
                "severity": "warning",
                "auto_resolution": "Will be imported as a payment/settlement record.",
            })

    return results


def _check_split(row: dict, row_num: int) -> list[dict]:
    results = []
    split_type = str(row.get("split_type", "")).strip().lower()
    split_details_raw = row.get("split_details")

    if split_type == "percentage" and split_details_raw:
        details = parse_split_details(str(split_details_raw))
        total = Decimal("0")
        for name, val in details.items():
            try:
                total += Decimal(val.replace("%", ""))
            except InvalidOperation:
                pass

        if total != 100:
            results.append({
                "row_number": row_num,
                "anomaly_type": "percentage_mismatch",
                "description": f"Percentages sum to {total}%, not 100%. Details: {split_details_raw}",
                "severity": "error",
                "auto_resolution": "Requires user to fix percentages to sum to 100%.",
            })

    if split_type == "unequal" and split_details_raw:
        details = parse_split_details(str(split_details_raw))
        try:
            amount = Decimal(str(row.get("amount", 0)))
            split_total = sum(Decimal(v) for v in details.values())
            if split_total != amount:
                results.append({
                    "row_number": row_num,
                    "anomaly_type": "split_mismatch",
                    "description": f"Unequal split amounts sum to {split_total} but expense is {amount}.",
                    "severity": "error",
                    "auto_resolution": "Requires user correction.",
                })
        except (InvalidOperation, ValueError):
            pass

    if split_type == "equal" and split_details_raw:
        results.append({
            "row_number": row_num,
            "anomaly_type": "split_mismatch",
            "description": f"Split type is 'equal' but split_details provided: '{split_details_raw}'. Ignoring details.",
            "severity": "info",
            "auto_resolution": "Using equal split, ignoring redundant split_details.",
        })

    return results


def _check_membership_temporal(row: dict, row_num: int, member_dates: dict) -> list[dict]:
    results = []
    date_val = row.get("date")
    if not date_val:
        return results

    try:
        if isinstance(date_val, str):
            expense_date = datetime.strptime(date_val, "%Y-%m-%d").date()
        elif isinstance(date_val, (datetime, date)):
            expense_date = date_val if isinstance(date_val, date) else date_val.date()
        else:
            return results
    except ValueError:
        return results

    participants = parse_participants(str(row.get("split_with", "")))
    for p in participants:
        if p in member_dates:
            info = member_dates[p]
            left = _to_date(info.get("left"))
            joined = _to_date(info.get("joined"))
            if left and expense_date > left:
                results.append({
                    "row_number": row_num,
                    "anomaly_type": "membership_violation",
                    "description": f"'{p}' left the group on {left} but is included in expense dated {expense_date}.",
                    "severity": "warning",
                    "auto_resolution": f"Flagged for user review. Consider removing {p} from this expense.",
                })
            if joined and expense_date < joined:
                results.append({
                    "row_number": row_num,
                    "anomaly_type": "membership_violation",
                    "description": f"'{p}' joined on {joined} but expense is dated {expense_date}.",
                    "severity": "warning",
                    "auto_resolution": f"Flagged for user review.",
                })

    return results


def _check_duplicate(row: dict, row_num: int, seen: list[tuple]) -> list[dict]:
    results = []
    desc = str(row.get("description", "")).strip().lower()
    amount = str(row.get("amount", ""))
    date_val = str(row.get("date", ""))
    paid_by = normalize_name(str(row.get("paid_by", "")))

    for prev_num, prev_row in seen:
        prev_desc = str(prev_row.get("description", "")).strip().lower()
        prev_amount = str(prev_row.get("amount", ""))
        prev_date = str(prev_row.get("date", ""))
        prev_paid_by = normalize_name(str(prev_row.get("paid_by", "")))

        if date_val == prev_date and paid_by == prev_paid_by:
            if amount == prev_amount and similarity(desc, prev_desc) > 0.6:
                results.append({
                    "row_number": row_num,
                    "anomaly_type": "duplicate",
                    "description": f"Likely duplicate of row {prev_num}. Same date, payer, amount. Descriptions: '{row.get('description')}' vs '{prev_row.get('description')}'",
                    "severity": "warning",
                    "auto_resolution": f"Recommend skipping this row (keeping row {prev_num}).",
                })
            elif similarity(desc, prev_desc) > 0.6 and amount != prev_amount:
                results.append({
                    "row_number": row_num,
                    "anomaly_type": "duplicate",
                    "description": f"Possible duplicate of row {prev_num} with different amount ({amount} vs {prev_amount}). '{row.get('description')}' vs '{prev_row.get('description')}'",
                    "severity": "error",
                    "auto_resolution": f"Conflicting amounts. User must choose which row to keep.",
                })

    return results


def _check_unknown_participants(row: dict, row_num: int, group_members: list[str]) -> list[dict]:
    results = []
    participants = parse_participants(str(row.get("split_with", "")))
    normalized_members = {m.lower(): m for m in group_members}

    for p in participants:
        if p.lower() not in normalized_members and p not in group_members:
            has_close_match = any(similarity(p, m) > 0.7 for m in group_members)
            if not has_close_match:
                results.append({
                    "row_number": row_num,
                    "anomaly_type": "unknown_participant",
                    "description": f"Unknown participant '{p}' in expense '{row.get('description')}'. Not a group member.",
                    "severity": "warning",
                    "auto_resolution": f"Will add '{p}' as a guest participant.",
                })

    return results
